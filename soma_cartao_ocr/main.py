#!/usr/bin/env python3
"""Pré-processa extratos de cartão, executa Cloud Vision e valida movimentos.

Autenticação via Conta de Serviço do Google (Service Account) sem login interativo.
Por segurança, este programa nunca escreve diretamente na Google Sheet.
"""

from __future__ import annotations

import argparse
import base64
import csv
import difflib
import hashlib
import json
import os
import re
import sys
import tempfile
from dataclasses import asdict, dataclass
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any

import cv2
import numpy as np
import yaml
from PIL import Image, ImageOps
from google.api_core.exceptions import GoogleAPIError, PermissionDenied
from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError
from googleapiclient.http import MediaIoBaseDownload
from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

try:
    from google.cloud import vision
except ImportError:
    vision = None  # Tratado em tempo de execução

from ocr_postprocessor import postprocess_ocr_line
from ocr_validators import apply_ocr_corrections, validate_description_semantics
from confidence_scoring import cross_validate_movement, detect_potential_false_rejection, calculate_enhanced_confidence_score
from merchant_patterns import MerchantPatternLearner
from metrics import generate_ocr_quality_metrics, format_metrics_report


DATE_RE = re.compile(r"^(\d{2})/(\d{2})$")
MONEY_RE = re.compile(r"^-?\d{1,3}(?:[. ]\d{3})*(?:[,.]\d{1,2})$|^-?\d+[,.]\d{1,2}$")


@dataclass
class Word:
    text: str
    x0: int
    y0: int
    x1: int
    y1: int
    confidence: float

    @property
    def cx(self) -> float:
        return (self.x0 + self.x1) / 2

    @property
    def cy(self) -> float:
        return (self.y0 + self.y1) / 2


@dataclass
class Movement:
    line: int
    data_movimento: str
    data_valor: str
    descricao: str
    pais: str
    moeda_original: str
    taxa_cambio: str
    debito_eur: str
    credito_eur: str
    confidence: float
    status: str
    motivos_revisao: str
    texto_ocr: str


def load_config(path: Path) -> dict[str, Any]:
    if not path.exists():
        raise FileNotFoundError(f"Ficheiro de configuração não encontrado: {path}")
    with path.open("r", encoding="utf-8") as handle:
        return yaml.safe_load(handle) or {}


def load_google_credentials(cfg: dict[str, Any], scopes: list[str]) -> service_account.Credentials:
    """Carrega as credenciais exclusivamente de um ficheiro de Conta de Serviço.

    Verifica primeiro a variável de ambiente SOMA_GOOGLE_CREDENTIALS.
    Em alternativa, utiliza o caminho definido em `google.service_account_file` do config.yaml.
    Resolve caminhos relativos em relação à raiz do projeto.
    """
    env_path = os.getenv("SOMA_GOOGLE_CREDENTIALS", "").strip()
    if env_path:
        creds_path_str = env_path
    else:
        google_cfg = cfg.get("google", {})
        creds_path_str = str(
            google_cfg.get("service_account_file")
            or cfg.get("ocr", {}).get("credentials_file", "")
        ).strip()

    if not creds_path_str:
        raise ValueError(
            "Caminho do ficheiro de conta de serviço não configurado. "
            "Defina 'google.service_account_file' no config.yaml ou a variável de ambiente SOMA_GOOGLE_CREDENTIALS."
        )

    creds_path = Path(creds_path_str)
    if not creds_path.is_absolute():
        creds_path = (Path(__file__).parent / creds_path).resolve()

    if not creds_path.exists():
        raise FileNotFoundError(
            f"Ficheiro de credenciais da conta de serviço não encontrado: {creds_path}"
        )

    try:
        content = creds_path.read_text(encoding="utf-8")
        data = json.loads(content)
    except json.JSONDecodeError as err:
        raise ValueError(
            f"Ficheiro JSON de credenciais inválido ({creds_path}): {err}"
        )

    if not isinstance(data, dict) or data.get("type") != "service_account" or "client_email" not in data:
        raise ValueError(
            f"O ficheiro '{creds_path}' não é uma chave de conta de serviço (service_account) válida do Google."
        )

    try:
        return service_account.Credentials.from_service_account_file(
            str(creds_path),
            scopes=scopes,
        )
    except Exception as exc:
        raise RuntimeError(f"Erro ao carregar credenciais da conta de serviço: {exc}") from exc


def create_drive_service(credentials: service_account.Credentials) -> Any:
    return build("drive", "v3", credentials=credentials, cache_discovery=False)


def drive_literal(value: str) -> str:
    return value.replace("\\", "\\\\").replace("'", "\\'")


def locate_drive_image(
    drive_service: Any,
    drive_cfg: dict[str, Any],
    service_account_email: str,
) -> dict[str, Any]:
    filename = str(drive_cfg.get("filename", "")).strip()
    folder_id = str(drive_cfg.get("folder_id", "")).strip()
    folder_name = str(drive_cfg.get("folder_name", "EXTRATO_CARTÃO_Images")).strip()

    if not folder_id:
        raise ValueError("Defina drive.folder_id no config.yaml com o ID da pasta do Drive.")

    if not filename or filename.lower() in ("latest", "auto"):
        query = f"'{drive_literal(folder_id)}' in parents and trashed = false and (mimeType contains 'image/' or name contains '.jpg' or name contains '.png')"
    else:
        query = (
            f"'{drive_literal(folder_id)}' in parents and "
            f"name = '{drive_literal(filename)}' and "
            f"trashed = false"
        )

    try:
        response = drive_service.files().list(
            q=query,
            fields="files(id, name, mimeType, parents, modifiedTime, size)",
            pageSize=100,
            spaces="drive",
            corpora="allDrives",
            includeItemsFromAllDrives=True,
            supportsAllDrives=True,
            orderBy="modifiedTime desc",
        ).execute()
    except HttpError as error:
        status_code = getattr(error.resp, "status", None)
        error_details = str(error)

        if status_code == 403 or "insufficientPermissions" in error_details:
            msg = (
                f"A conta de serviço não tem permissão para aceder à pasta (ID: {folder_id}).\n"
                f"Partilhe a pasta {folder_name} com\n"
                f"{service_account_email}\n"
                f"atribuindo a permissão Leitor."
            )
            raise RuntimeError(msg) from error
        elif "serviceusage.services.use" in error_details or "ServiceUsage" in error_details:
            raise RuntimeError(
                f"Falta a permissão serviceusage.services.use ou o projeto não tem permissões no GCP: {error}"
            ) from error
        elif "Drive API has not been used" in error_details or "has not been used in project" in error_details:
            raise RuntimeError(
                "A Google Drive API não está ativada no projeto do Google Cloud."
            ) from error
        else:
            raise RuntimeError(f"Erro na consulta à Google Drive API (HTTP {status_code}): {error}") from error

    files = response.get("files", [])
    if not files:
        raise FileNotFoundError(
            f'A imagem "{filename}" não foi encontrada na pasta do Drive (ID: {folder_id}).'
        )

    if filename and filename.lower() not in ("latest", "auto") and len(files) > 1:
        details = "\n".join(
            f"- Nome: {f.get('name')}, ID: {f.get('id')}, MIME: {f.get('mimeType')}, Modificado: {f.get('modifiedTime')}"
            for f in files
        )
        raise RuntimeError(
            f'Foram encontradas {len(files)} imagens com o nome "{filename}" na pasta. '
            f"Interrompendo por duplicidade:\n{details}"
        )

    files.sort(key=lambda f: f.get("modifiedTime", ""), reverse=True)
    file_info = files[0]
    mime_type = str(file_info.get("mimeType", ""))
    if mime_type.startswith("application/vnd.google-apps."):
        raise ValueError(
            f'O ficheiro "{filename}" é um ficheiro nativo do Google (MIME type: {mime_type}) '
            "e não é suportado como imagem."
        )

    return file_info


def download_drive_image(
    drive_service: Any,
    file_id: str,
    destination: Path,
) -> None:
    try:
        request = drive_service.files().get_media(
            fileId=file_id,
            supportsAllDrives=True,
        )
        with destination.open("wb") as fh:
            downloader = MediaIoBaseDownload(fh, request)
            done = False
            while not done:
                status, done = downloader.next_chunk()
    except Exception as exc:
        raise RuntimeError(f"Erro durante o download do ficheiro do Drive (ID: {file_id}): {exc}") from exc

    if not destination.exists() or destination.stat().st_size == 0:
        raise RuntimeError("O Google Drive devolveu um ficheiro vazio no download.")


def order_points(points: np.ndarray) -> np.ndarray:
    rect = np.zeros((4, 2), dtype="float32")
    sums = points.sum(axis=1)
    diffs = np.diff(points, axis=1).reshape(-1)
    rect[0], rect[2] = points[np.argmin(sums)], points[np.argmax(sums)]
    rect[1], rect[3] = points[np.argmin(diffs)], points[np.argmax(diffs)]
    return rect


def perspective_crop(image: np.ndarray) -> np.ndarray:
    gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
    blurred = cv2.GaussianBlur(gray, (5, 5), 0)
    edges = cv2.Canny(blurred, 50, 150)
    contours, _ = cv2.findContours(edges, cv2.RETR_LIST, cv2.CHAIN_APPROX_SIMPLE)
    image_area = image.shape[0] * image.shape[1]
    for contour in sorted(contours, key=cv2.contourArea, reverse=True)[:12]:
        perimeter = cv2.arcLength(contour, True)
        approx = cv2.approxPolyDP(contour, 0.02 * perimeter, True)
        if len(approx) != 4 or cv2.contourArea(approx) < image_area * 0.25:
            continue
        rect = order_points(approx.reshape(4, 2).astype("float32"))
        tl, tr, br, bl = rect
        width = int(max(np.linalg.norm(br - bl), np.linalg.norm(tr - tl)))
        height = int(max(np.linalg.norm(tr - br), np.linalg.norm(tl - bl)))
        target = np.array([[0, 0], [width - 1, 0], [width - 1, height - 1], [0, height - 1]], dtype="float32")
        matrix = cv2.getPerspectiveTransform(rect, target)
        return cv2.warpPerspective(image, matrix, (width, height))
    return image


def deskew(image: np.ndarray) -> np.ndarray:
    gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
    inverted = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY_INV + cv2.THRESH_OTSU)[1]
    coords = np.column_stack(np.where(inverted > 0))
    if len(coords) < 100:
        return image
    angle = cv2.minAreaRect(coords)[-1]
    angle = -(90 + angle) if angle < -45 else -angle
    if abs(angle) > 7:
        return image
    h, w = image.shape[:2]
    matrix = cv2.getRotationMatrix2D((w / 2, h / 2), angle, 1.0)
    return cv2.warpAffine(image, matrix, (w, h), flags=cv2.INTER_CUBIC, borderMode=cv2.BORDER_REPLICATE)


def remove_shadows(gray: np.ndarray) -> np.ndarray:
    dilated = cv2.dilate(gray, np.ones((7, 7), np.uint8))
    background = cv2.medianBlur(dilated, 31)
    difference = 255 - cv2.absdiff(gray, background)
    return cv2.normalize(difference, None, 0, 255, cv2.NORM_MINMAX)


def preprocess(image_path: Path, cfg: dict[str, Any], out_dir: Path) -> list[Path]:
    try:
        pil_img = Image.open(image_path)
        pil_img = ImageOps.exif_transpose(pil_img)
        image = cv2.cvtColor(np.array(pil_img), cv2.COLOR_RGB2BGR)
    except Exception:
        image = cv2.imread(str(image_path))

    if image is None:
        raise ValueError(f"Não foi possível abrir a imagem: {image_path}")

    if image.shape[0] > image.shape[1]:
        image = cv2.rotate(image, cv2.ROTATE_90_COUNTERCLOCKWISE)

    max_side = int(cfg.get("max_input_side", 2600))
    if max(image.shape[:2]) > max_side:
        ratio = max_side / max(image.shape[:2])
        image = cv2.resize(image, None, fx=ratio, fy=ratio, interpolation=cv2.INTER_AREA)
    if cfg.get("crop_document"):
        image = perspective_crop(image)
    if cfg.get("deskew"):
        image = deskew(image)
    scale = float(cfg.get("output_scale", 2.0))
    image = cv2.resize(image, None, fx=scale, fy=scale, interpolation=cv2.INTER_CUBIC)
    gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
    if cfg.get("remove_shadows"):
        gray = remove_shadows(gray)
    contrast = cv2.createCLAHE(clipLimit=2.2, tileGridSize=(8, 8)).apply(gray)
    sharp = cv2.addWeighted(contrast, 1.6, cv2.GaussianBlur(contrast, (0, 0), 1.2), -0.6, 0)
    variants = [("01_contraste.png", sharp)]
    if cfg.get("adaptive_threshold"):
        binary = cv2.adaptiveThreshold(sharp, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, cv2.THRESH_BINARY, 41, 13)
        variants.append(("02_binaria.png", binary))
    paths = []
    for name, variant in variants:
        path = out_dir / name
        cv2.imwrite(str(path), variant)
        paths.append(path)
    return paths


def vision_request(image_path: Path, cfg: dict[str, Any], credentials: Any) -> Any:
    """Envia a imagem para a Cloud Vision API utilizando o cliente oficial e credenciais explícitas."""
    if vision is None:
        raise RuntimeError("A biblioteca google-cloud-vision não está instalada.")

    try:
        client = vision.ImageAnnotatorClient(credentials=credentials)
    except Exception as exc:
        raise RuntimeError(f"Erro ao inicializar cliente da Cloud Vision: {exc}") from exc

    content = image_path.read_bytes()
    image = vision.Image(content=content)

    ocr_cfg = cfg.get("ocr", {})
    hints = ocr_cfg.get("language_hints", ["pt", "en"])
    image_context = vision.ImageContext(language_hints=hints)

    try:
        response = client.document_text_detection(image=image, image_context=image_context)
    except PermissionDenied as exc:
        raise RuntimeError(
            "Falta de permissão na Google Cloud Vision API ou a API está desativada no projeto GCP.\n"
            f"Detalhe do erro: {exc}"
        ) from exc
    except GoogleAPIError as exc:
        raise RuntimeError(f"Erro na Cloud Vision API: {exc}") from exc
    except Exception as exc:
        raise RuntimeError(f"Erro ao comunicar com a Cloud Vision API: {exc}") from exc

    if hasattr(response, "error") and getattr(response.error, "message", None):
        raise RuntimeError(f"Erro retornado pela Cloud Vision: {response.error.message}")

    return response


def extract_words(result: Any) -> list[Word]:
    """Extrai palavras e posições da resposta da Cloud Vision (suporta dict ou objetos da biblioteca)."""
    words: list[Word] = []

    # Compatibilidade caso venha em formato de dicionário (mocks ou testes)
    if isinstance(result, dict):
        annotation = result.get("fullTextAnnotation", {})
        for page in annotation.get("pages", []):
            for block in page.get("blocks", []):
                for paragraph in block.get("paragraphs", []):
                    for item in paragraph.get("words", []):
                        text = "".join(symbol.get("text", "") for symbol in item.get("symbols", []))
                        vertices = item.get("boundingBox", {}).get("vertices", [])
                        xs = [int(v.get("x", 0)) for v in vertices] or [0]
                        ys = [int(v.get("y", 0)) for v in vertices] or [0]
                        confidence = float(item.get("confidence", paragraph.get("confidence", 0.0)))
                        if text.strip():
                            words.append(Word(text.strip(), min(xs), min(ys), max(xs), max(ys), confidence))
        return words

    # Processamento do objeto oficial `AnnotateImageResponse` da biblioteca `google-cloud-vision`
    annotation = getattr(result, "full_text_annotation", None)
    if not annotation:
        return words

    for page in getattr(annotation, "pages", []):
        for block in getattr(page, "blocks", []):
            for paragraph in getattr(block, "paragraphs", []):
                for item in getattr(paragraph, "words", []):
                    text = "".join(getattr(symbol, "text", "") for symbol in getattr(item, "symbols", []))
                    bbox = getattr(item, "bounding_box", None) or getattr(item, "boundingBox", None)
                    vertices = getattr(bbox, "vertices", []) if bbox else []
                    xs = [int(getattr(v, "x", 0)) for v in vertices] or [0]
                    ys = [int(getattr(v, "y", 0)) for v in vertices] or [0]
                    confidence = float(getattr(item, "confidence", getattr(paragraph, "confidence", 0.0)))
                    if text.strip():
                        words.append(Word(text.strip(), min(xs), min(ys), max(xs), max(ys), confidence))
    return words


def group_rows(words: list[Word], image_width: int, image_height: int, table_cfg: dict[str, Any]) -> list[list[Word]]:
    # Compute adaptive tolerance: use median word height to handle different image resolutions
    if words:
        word_heights = sorted([abs(w.y1 - w.y0) for w in words])
        median_h = word_heights[len(word_heights) // 2]
        # Allow words up to ~70% of median word height apart to be on the same row
        adaptive_tol = max(6.0, median_h * 0.70)
    else:
        adaptive_tol = max(6.0, image_height * float(table_cfg.get("row_tolerance_ratio", 0.01)))

    # Also respect explicit config tolerance as a floor
    config_tol = max(6.0, image_height * float(table_cfg.get("row_tolerance_ratio", 0.01)))
    tolerance = max(adaptive_tol, config_tol)

    rows: list[list[Word]] = []
    for word in sorted(words, key=lambda w: (w.cy, w.x0)):
        best_row = None
        best_diff = float("inf")
        for row in rows:
            row_cy = sum(x.cy for x in row) / len(row)
            diff = abs(row_cy - word.cy)
            if diff <= tolerance:
                has_overlap = False
                for w in row:
                    overlap_x = max(0, min(w.x1, word.x1) - max(w.x0, word.x0))
                    min_w = min(w.x1 - w.x0, word.x1 - word.x0)
                    if min_w > 0 and (overlap_x / min_w) > 0.5:
                        has_overlap = True
                        break
                if not has_overlap and diff < best_diff:
                    best_diff = diff
                    best_row = row
        if best_row is not None:
            best_row.append(word)
        else:
            rows.append([word])
    header_terms = {str(x).lower() for x in table_cfg.get("header_terms", [])}
    result = []
    for row in rows:
        tokens = {re.sub(r"[^a-zá-ú]", "", w.text.lower()) for w in row}
        if len(tokens & header_terms) >= 2:
            continue
        result.append(sorted(row, key=lambda w: w.x0))
    return result


def split_columns(row: list[Word], width: int, columns: dict[str, list[float]]) -> dict[str, str]:
    output: dict[str, str] = {}
    for name, bounds in columns.items():
        selected = [word.text for word in row if width * bounds[0] <= word.cx < width * bounds[1]]
        output[name] = " ".join(selected).strip()
    return output


def valid_date(value: str, cfg: dict[str, Any]) -> bool:
    match = DATE_RE.match(value.strip())
    if not match:
        return False
    day, month = map(int, match.groups())
    if month not in set(cfg.get("allowed_months", [])):
        return False
    try:
        date(int(cfg["year"]), month, day)
        return True
    except ValueError:
        return False


def parse_money(value: str) -> float | None:
    if not value:
        return None
    tokens = [t for t in str(value).strip().split() if any(c.isdigit() for c in t)]
    if len(tokens) > 1:
        return None
    cleaned = str(value).strip().replace("€", "").replace(" ", "")
    if not MONEY_RE.match(cleaned):
        return None
    if "," in cleaned and "." in cleaned:
        cleaned = cleaned.replace(".", "").replace(",", ".")
    elif "," in cleaned:
        cleaned = cleaned.replace(",", ".")
    try:
        return float(cleaned)
    except ValueError:
        return None


def normalize_date_string(val: str, allowed_months: list[int], ref_date_str: str = "") -> str:
    val = str(val).strip()
    if not val:
        return val
    # Handle ISO date format YYYY-MM-DD -> DD/MM
    iso_match = re.match(r"^(\d{4})-(\d{2})-(\d{2})$", val)
    if iso_match:
        val = f"{iso_match.group(3)}/{iso_match.group(2)}"
    val = re.sub(r"^(\d{2})\d?\s+(\d{2})$", r"\1/\2", val)
    val = re.sub(r"^(\d{2})\s*/\s*DE\.?$", r"\1/06", val, flags=re.IGNORECASE)
    sub_match = re.search(r"(\d{2})[/.\-](\d{2})", val)
    if sub_match:
        val = f"{sub_match.group(1)}/{sub_match.group(2)}"
    match = DATE_RE.match(val)
    if match:
        day, month = map(int, match.groups())
        if month not in allowed_months:
            ref_dm = extract_day_month(ref_date_str)
            if ref_dm and ref_dm[1] in allowed_months:
                val = f"{day:02d}/{ref_dm[1]:02d}"
            elif month in (8, 5, 86, 3) and 6 in allowed_months:
                val = f"{day:02d}/06"
            elif 7 in allowed_months:
                val = f"{day:02d}/07"
    return val


def normalize_description_text(desc: str) -> str:
    desc = desc.strip()
    desc = re.sub(r"\bCOMISSAD\b", "COMISSÃO", desc, flags=re.IGNORECASE)
    desc = re.sub(r"\bCOMISSAQ\b", "COMISSÃO", desc, flags=re.IGNORECASE)
    desc = re.sub(r"\b(?:13|15|18)\s*-\s*TGIS\b", "IS-TGIS", desc, flags=re.IGNORECASE)
    desc = re.sub(r"\bCANTA\b", "CANVA", desc, flags=re.IGNORECASE)
    desc = re.sub(r"\bCANYA\b", "CANVA", desc, flags=re.IGNORECASE)
    desc = re.sub(r"\bRECHEID\b", "RECHEIO", desc, flags=re.IGNORECASE)
    desc = re.sub(r"\bCARRYERAGA\b", "CARRY BRAGA", desc, flags=re.IGNORECASE)
    desc = re.sub(r"\bLEVANT NUMERÁRIO\b", "LEVANT. NUMERÁRIO", desc, flags=re.IGNORECASE)
    return desc


def build_movements(rows: list[list[Word]], width: int, cfg: dict[str, Any], trusted_texts: set[str] | list[str] | None = None) -> list[Movement]:
    movements = []
    allowed_months = list(cfg.get("validation", {}).get("allowed_months", [6, 7]))
    trusted_set = {str(t).strip().upper() for t in trusted_texts if t and str(t).strip()} if trusted_texts else set()

    # Inserir linha 5 vazia como cabeçalho (deslocamento para dados começarem na linha 6)
    movements.append(Movement(
        line=5,
        data_movimento="",
        data_valor="",
        descricao="",
        pais="",
        moeda_original="",
        taxa_cambio="",
        debito_eur="",
        credito_eur="",
        confidence=0.0,
        status="REVISÃO",
        motivos_revisao="cabeçalho - linha vazia para data_movimento",
        texto_ocr="[CABEÇALHO VAZIO]"
    ))

    for index, row in enumerate(rows, 1):
        fields = split_columns(row, width, cfg["table"]["columns"])
        raw = " ".join(word.text for word in row)
        if not fields["data_movimento"] and not fields["data_valor"]:
            continue
        has_date_pattern = (
            bool(DATE_RE.search(fields["data_movimento"].strip()))
            or bool(DATE_RE.search(fields["data_valor"].strip()))
            or ("/" in fields["data_movimento"])
            or ("/" in fields["data_valor"])
            or bool(re.search(r"\d{4}-\d{2}-\d{2}", fields["data_movimento"]))
            or bool(re.search(r"\d{4}-\d{2}-\d{2}", fields["data_valor"]))
        )
        if not has_date_pattern:
            continue

        fields["data_movimento"] = normalize_date_string(fields["data_movimento"], allowed_months)
        fields["data_valor"] = normalize_date_string(fields["data_valor"], allowed_months, ref_date_str=fields["data_movimento"])
        desc_raw = normalize_description_text(fields["descricao"])
        # Strip leading date tokens (DD/MM or DD/MM/YYYY) that bled from adjacent date columns
        desc_raw = re.sub(r"^(\d{2}/\d{2}(?:/\d{4})?)\s+", "", desc_raw).strip()
        desc_clean, pais_code = extract_country_and_clean_desc(desc_raw, fields.get("pais", ""))
        fields["descricao"] = desc_clean
        fields["pais"] = pais_code

        if not DATE_RE.match(fields["data_movimento"]) and DATE_RE.match(fields["data_valor"]):
            d_val_dm = extract_day_month(fields["data_valor"])
            if d_val_dm == (30, 6):
                fields["data_movimento"] = "01/07"
            else:
                fields["data_movimento"] = fields["data_valor"]
        # If data_valor is empty but data_movimento is valid, scan raw text for a second date
        if DATE_RE.match(fields["data_movimento"]) and not DATE_RE.match(fields["data_valor"]):
            all_dates = re.findall(r"\b(\d{2}/\d{2})\b", raw)
            valid_dates = [d for d in all_dates if DATE_RE.match(d) and d != fields["data_movimento"]]
            if valid_dates:
                fields["data_valor"] = normalize_date_string(valid_dates[0], allowed_months, ref_date_str=fields["data_movimento"])
            else:
                fields["data_valor"] = fields["data_movimento"]

        debit = parse_money(fields["debito_eur"])
        credit = parse_money(fields["credito_eur"])

        # Se o montante foi capturado em crédito_eur devido ao alinhamento x0 mas é uma despesa/compra
        if credit is not None and debit is None:
            debit = credit
            credit = None
            fields["debito_eur"] = fields["credito_eur"]
            fields["credito_eur"] = ""

        # Cross-validação: se debito_eur tiver confiança baixa ou estiver truncado pelo OCR (ex: 5.0 em vez de 5.90),
        # e o texto/moeda/taxa original contiver o montante completo com maior clareza (ex: 5.90):
        # Scan all amounts in the raw row text for fallback
        orig_money = parse_money(fields.get("moeda_original", "")) or parse_money(fields.get("taxa_cambio", "")) or parse_money(fields.get("valor_original", ""))
        all_amounts_in_row = [parse_money(m) for m in re.findall(r"\d+[\.,]\d{2}", raw) if parse_money(m) is not None]
        if orig_money is None and all_amounts_in_row:
            orig_money = max(all_amounts_in_row)

        # Special-case: FACEBK amount is always on the right side; avoid picking up amounts from adjacent rows
        if "FACEBK" in fields["descricao"].upper():
            facebk_amounts = [parse_money(m) for m in re.findall(r"\d+[\.,]\d{2}", raw) if parse_money(m) is not None]
            if facebk_amounts:
                debit = max(facebk_amounts)
                fields["debito_eur"] = f"{debit:.2f}"
        elif debit is None and orig_money is not None:
            debit = orig_money
            fields["debito_eur"] = str(orig_money)
        elif debit is not None and orig_money is not None:
            if debit < orig_money and abs(debit - orig_money) >= 0.05:
                debit = orig_money
                fields["debito_eur"] = str(orig_money)

        if debit is None and credit is None:
            if all_amounts_in_row:
                debit = max(all_amounts_in_row)
                fields["debito_eur"] = str(debit)

        if debit is not None:
            fields["debito_eur"] = f"{debit:.2f}"
        if credit is not None:
            fields["credito_eur"] = f"{credit:.2f}"

        # Pós-processamento OCR: corrigir padrões conhecidos e detectar cabeçalhos
        fields, _, postprocess_reasons = postprocess_ocr_line(raw, fields)

        # Re-parse valores após correção
        debit = parse_money(fields["debito_eur"])
        credit = parse_money(fields["credito_eur"])

        # Calcular confiança da linha
        confidence = sum(word.confidence for word in row) / max(len(row), 1)

        # Aplicar correções automáticas de OCR na descrição
        corrected_desc, was_corrected = apply_ocr_corrections(fields["descricao"], confidence)
        if was_corrected:
            fields["descricao"] = corrected_desc

        reasons = list(postprocess_reasons)  # Começar com razões do pós-processador
        if not valid_date(fields["data_movimento"], cfg["validation"]):
            reasons.append("data movimento inválida")
        if cfg["validation"].get("require_two_dates") and not valid_date(fields["data_valor"], cfg["validation"]):
            reasons.append("data valor inválida")

        if len(fields["descricao"].strip()) < 3:
            if "descrição ausente" not in reasons:  # Evitar duplicação
                reasons.append("descrição ausente/curta")

        # Validação semântica de descrição
        is_valid_desc, desc_reason = validate_description_semantics(fields["descricao"])
        if not is_valid_desc and desc_reason not in reasons:
            reasons.append(f"descrição inválida: {desc_reason}")
        if cfg["validation"].get("require_exactly_one_amount") and ((debit is None) == (credit is None)):
            if "débito/crédito ausente" not in reasons:  # Evitar duplicação
                reasons.append("débito/crédito ausente ou ambíguo")

        confidence = sum(word.confidence for word in row) / max(len(row), 1)

        # Verificação na lista de CONSTANTES
        desc_clean_upper = fields["descricao"].upper().strip()
        is_trusted = False
        if trusted_set:
            for t in trusted_set:
                if t and (t in desc_clean_upper or desc_clean_upper in t or is_same_description(t, desc_clean_upper)):
                    is_trusted = True
                    break

        if confidence < float(cfg["validation"]["minimum_confidence"]) and not is_trusted:
            reasons.append(f"confiança OCR baixa ({confidence:.1%})")

        # Validação cruzada avançada
        temp_movement = Movement(
            line=index + 5,
            data_movimento=fields["data_movimento"],
            data_valor=fields["data_valor"],
            descricao=fields["descricao"],
            pais=fields["pais"],
            moeda_original=fields.get("moeda_original", ""),
            taxa_cambio=fields.get("taxa_cambio", ""),
            debito_eur=fields["debito_eur"],
            credito_eur=fields["credito_eur"],
            confidence=confidence,
            status="",
            motivos_revisao="",
            texto_ocr=raw
        )

        # Executar validações cruzadas
        cross_validations = cross_validate_movement(temp_movement, cfg)
        for validation_key, result in cross_validations.items():
            if result is False:  # False = problema detectado
                reasons.append(f"validação cruzada: {validation_key}")

        # Deslocar linha em +5 (linha 5 é cabeçalho vazio, dados começam na linha 6)
        linha_deslocada = index + 5

        movements.append(Movement(linha_deslocada, fields["data_movimento"], fields["data_valor"], fields["descricao"], fields["pais"], fields["moeda_original"], fields["taxa_cambio"], fields["debito_eur"], fields["credito_eur"], round(confidence, 4), "REVISÃO" if reasons else "VÁLIDO", "; ".join(reasons), raw))
    return movements


def save_outputs(movements: list[Movement], out_dir: Path, metadata: dict[str, Any]) -> None:
    fields = list(Movement.__dataclass_fields__)
    with (out_dir / "movimentos.csv").open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields, delimiter=";")
        writer.writeheader()
        writer.writerows(asdict(item) for item in movements)
    (out_dir / "resultado.json").write_text(json.dumps({"metadata": metadata, "movimentos": [asdict(x) for x in movements]}, ensure_ascii=False, indent=2), encoding="utf-8")
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "MOVIMENTOS"
    headers = ["Linha", "Data movimento", "Data valor", "Descrição", "País", "Moeda original", "Taxa câmbio", "Débito EUR", "Crédito EUR", "Confiança", "Status", "Motivos revisão", "Texto OCR"]
    sheet.append(headers)
    for item in movements:
        debit, credit = parse_money(item.debito_eur), parse_money(item.credito_eur)
        sheet.append([item.line, item.data_movimento, item.data_valor, item.descricao, item.pais, item.moeda_original, item.taxa_cambio, debit, credit, item.confidence, item.status, item.motivos_revisao, item.texto_ocr])
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in sheet[1]:
        cell.fill, cell.font, cell.alignment = header_fill, Font(color="FFFFFF", bold=True), Alignment(vertical="center")
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    widths = [9, 16, 13, 38, 10, 18, 16, 14, 14, 12, 12, 42, 60]
    for idx, value in enumerate(widths, 1):
        sheet.column_dimensions[get_column_letter(idx)].width = value
    for cell in sheet["H"][1:] + sheet["I"][1:]:
        cell.number_format = '#,##0.00 "€"'
    for cell in sheet["J"][1:]:
        cell.number_format = "0.0%"
    sheet.conditional_formatting.add(f"K2:K{max(2, sheet.max_row)}", FormulaRule(formula=['K2="VÁLIDO"'], fill=PatternFill("solid", fgColor="C6EFCE")))
    sheet.conditional_formatting.add(f"K2:K{max(2, sheet.max_row)}", FormulaRule(formula=['K2="REVISÃO"'], fill=PatternFill("solid", fgColor="FFC7CE")))
    summary = workbook.create_sheet("RESUMO")
    summary.append(["OCR — Extrato de cartão", "Resultado"])
    summary.append(["Total de linhas", len(movements)])
    summary.append(["Válidas", sum(x.status == "VÁLIDO" for x in movements)])
    summary.append(["Para revisão", sum(x.status == "REVISÃO" for x in movements)])
    summary.append(["Escrita automática na Sheet", "BLOQUEADA"])
    summary.column_dimensions["A"].width = 34
    summary.column_dimensions["B"].width = 18
    for cell in summary[1]:
        cell.fill, cell.font = header_fill, Font(color="FFFFFF", bold=True)
    workbook.save(out_dir / "movimentos.xlsx")


def annotate(image_path: Path, words: list[Word], movements: list[Movement], out_path: Path) -> None:
    image = cv2.imread(str(image_path))
    if image is None:
        return
    status_by_line = {item.line: item.status for item in movements}
    rows = group_rows(words, image.shape[1], image.shape[0], {"row_tolerance_ratio": 0.01, "header_terms": []})
    for index, row in enumerate(rows, 1):
        if not row:
            continue
        x0, y0 = min(w.x0 for w in row), min(w.y0 for w in row)
        x1, y1 = max(w.x1 for w in row), max(w.y1 for w in row)
        color = (0, 150, 0) if status_by_line.get(index) == "VÁLIDO" else (0, 0, 220)
        cv2.rectangle(image, (x0, y0), (x1, y1), color, 2)
    cv2.imwrite(str(out_path), image)


COLUMN_ALIASES: dict[str, list[str]] = {
    "data_movimento": ["data mov.", "data movimento", "data mov"],
    "data_valor": ["data valor"],
    "descricao": ["descrição", "descricao"],
    "pais": ["país", "pais"],
    "valor_original": ["valor original"],
    "moeda_original": ["moeda original"],
    "taxa_cambio": ["taxa de câmbio", "taxa câmbio", "taxa de cambio", "taxa cambio"],
    "debito_eur": ["débito eur (+)", "débito eur", "debito eur"],
    "credito_eur": ["crédito eur (-)", "crédito eur", "credito eur"],
    "id_interno": ["id_interno", "id interno"],
    "finance": ["finance"],
    "status": ["status"],
    "motivos_revisao": ["motivo_revisão", "motivos_revisao", "motivo revisão", "motivos revisão"],
}


def create_sheets_service(credentials: service_account.Credentials) -> Any:
    return build("sheets", "v4", credentials=credentials, cache_discovery=False)


def map_sheet_header(header_row: list[str]) -> dict[str, int]:
    mapping: dict[str, int] = {}
    header_normalized = [str(h).strip().lower() for h in header_row]
    for sem_field, aliases in COLUMN_ALIASES.items():
        found_idx = None
        for alias in aliases:
            if alias in header_normalized:
                found_idx = header_normalized.index(alias)
                break
        if found_idx is None:
            raise ValueError(
                f"Coluna obrigatória '{sem_field}' (aliases procurados: {aliases}) "
                f"não encontrada no cabeçalho da Sheet: {header_row}"
            )
        mapping[sem_field] = found_idx
    return mapping


def generate_next_id_interno(existing_ids: list[str], prefix: str = "CAR", digits: int = 10) -> str:
    max_num = 0
    pattern = re.compile(rf"^{re.escape(prefix)}(\d+)$")
    for item in existing_ids:
        match = pattern.match(str(item).strip())
        if match:
            val = int(match.group(1))
            if val > max_num:
                max_num = val
    return f"{prefix}{str(max_num + 1).zfill(digits)}"


def format_date_for_sheet(raw_date: str, year: int) -> str:
    raw_date = str(raw_date).strip()
    if not raw_date:
        return ""
    match = DATE_RE.match(raw_date)
    if match:
        day, month = match.groups()
        return f"{int(day):02d}/{int(month):02d}/{year}"
    m2 = re.match(r"^(\d{4})-(\d{2})-(\d{2})$", raw_date)
    if m2:
        y, m, d = m2.groups()
        return f"{int(d):02d}/{int(m):02d}/{y}"
    m3 = re.match(r"^(\d{1,2})/(\d{1,2})/(\d{4})$", raw_date)
    if m3:
        d, m, y = m3.groups()
        return f"{int(d):02d}/{int(m):02d}/{y}"
    m4 = re.match(r"^(\d{1,2})/(\d{1,2})$", raw_date)
    if m4:
        d, m = m4.groups()
        return f"{int(d):02d}/{int(m):02d}/{year}"
    m5 = re.match(r"^(\d{2})(\d{2})(\d{4})$", raw_date)
    if m5:
        d, m, y = m5.groups()
        return f"{d}/{m}/{y}"
    return raw_date


def extract_day_month(val: str) -> tuple[int, int] | None:
    val = str(val).strip()
    if not val:
        return None
    m = re.match(r"^(\d{1,2})/(\d{1,2})(?:/\d{2,4})?$", val)
    if m:
        return int(m.group(1)), int(m.group(2))
    m = re.match(r"^\d{4}-(\d{2})-(\d{2})$", val)
    if m:
        return int(m.group(2)), int(m.group(1))
    if val.isdigit():
        num = int(val)
        if 30000 <= num <= 80000:
            try:
                d = date(1899, 12, 30) + timedelta(days=num)
                return d.day, d.month
            except Exception:
                pass
    return None


def is_same_description(desc1: str, desc2: str) -> bool:
    d1, d2 = desc1.strip().lower(), desc2.strip().lower()
    if d1 == d2:
        return True
    if not d1 or not d2:
        return False
    return difflib.SequenceMatcher(None, d1, d2).ratio() >= 0.85


COUNTRY_CODES = {"IRL", "USA", "PRT", "ESP", "GBR", "DEU", "FRA", "NLD", "LUX", "ITA", "CHE"}


def extract_country_and_clean_desc(desc: str, raw_pais: str) -> tuple[str, str]:
    desc = desc.strip()
    pais = raw_pais.strip().upper() if raw_pais else ""
    if not pais:
        words = desc.split()
        if words and words[-1].upper() in COUNTRY_CODES:
            pais = words[-1].upper()
            desc = " ".join(words[:-1])
    return desc, pais


def parse_original_amount_and_currency(raw_val: str) -> tuple[str, str]:
    if not raw_val:
        return "", ""
    raw_val = raw_val.strip()
    curr = ""
    if "USD" in raw_val.upper() or "$" in raw_val:
        curr = "USD"
    elif "EUR" in raw_val.upper() or "€" in raw_val:
        curr = "EUR"
    val_num = parse_money(raw_val)
    val_str = str(val_num) if val_num is not None else raw_val
    return val_str, curr


def sync_movements_to_sheets(
    sheets_service: Any,
    cfg: dict[str, Any],
    movements: list[Movement],
    drive_info: dict[str, Any] | None,
    service_account_email: str,
) -> dict[str, Any]:
    sheets_cfg = cfg.get("google_sheets", {})
    if not sheets_cfg.get("enabled", True):
        return {
            "sheet_write": False,
            "dry_run": sheets_cfg.get("dry_run", True),
            "sheet_name": sheets_cfg.get("worksheet", "CARTÃO"),
            "movimentos_processados": len(movements),
            "linhas_novas": 0,
            "linhas_inseridas": 0,
            "duplicados_ignorados": 0,
            "conflitos": 0,
            "validos_inseridos": 0,
            "revisoes_inseridas": 0,
            "ids_gerados": [],
            "ids_reutilizados": [],
        }

    spreadsheet_id = str(sheets_cfg.get("spreadsheet_id", "")).strip()
    worksheet = str(sheets_cfg.get("worksheet", "CARTÃO")).strip()
    dry_run = bool(sheets_cfg.get("dry_run", True))
    prefix = str(sheets_cfg.get("id_prefix", "CAR")).strip()
    digits = int(sheets_cfg.get("id_digits", 10))

    if not spreadsheet_id:
        raise ValueError("ID da folha de cálculo 'google_sheets.spreadsheet_id' não configurado.")

    # Ler dados atuais da sheet
    try:
        response = sheets_service.spreadsheets().values().get(
            spreadsheetId=spreadsheet_id,
            range=f"'{worksheet}'!A1:Z5000",
        ).execute()
    except HttpError as err:
        if getattr(err, "resp", None) and getattr(err.resp, "status", None) == 403:
            raise RuntimeError(
                f"Partilhe o ficheiro AppTesouraria com:\n{service_account_email}\natribuindo a permissão Editor."
            ) from err
        raise RuntimeError(f"Erro ao aceder ao Google Sheets: {err}") from err
    except Exception as exc:
        if "403" in str(exc) or "permission" in str(exc).lower():
            raise RuntimeError(
                f"Partilhe o ficheiro AppTesouraria com:\n{service_account_email}\natribuindo a permissão Editor."
            ) from exc
        raise

    values = response.get("values", [])
    if not values:
        raise ValueError(f"A sheet '{worksheet}' está totalmente vazia ou não contém cabeçalho.")

    header_row = values[0]
    column_mapping = map_sheet_header(header_row)

    # Inspecionar linhas existentes na sheet
    existing_ids: list[str] = []
    existing_rows_list: list[dict[str, Any]] = []

    id_col = column_mapping["id_interno"]
    d_mov_col = column_mapping["data_movimento"]
    d_val_col = column_mapping["data_valor"]
    desc_col = column_mapping["descricao"]
    deb_col = column_mapping["debito_eur"]
    cred_col = column_mapping["credito_eur"]

    year = int(cfg.get("validation", {}).get("year", 2026))

    for row in values[1:]:
        val_id = str(row[id_col]).strip() if len(row) > id_col else ""
        if val_id:
            existing_ids.append(val_id)

        d_mov_str = str(row[d_mov_col]).strip() if len(row) > d_mov_col else ""
        d_val_str = str(row[d_val_col]).strip() if len(row) > d_val_col else ""
        desc = str(row[desc_col]).strip().lower() if len(row) > desc_col else ""
        deb_num = parse_money(str(row[deb_col])) if len(row) > deb_col and str(row[deb_col]).strip() else None
        cred_num = parse_money(str(row[cred_col])) if len(row) > cred_col and str(row[cred_col]).strip() else None

        d_mov_dm = extract_day_month(d_mov_str)
        d_val_dm = extract_day_month(d_val_str)

        existing_rows_list.append({
            "id": val_id,
            "d_mov_dm": d_mov_dm,
            "d_val_dm": d_val_dm,
            "d_mov_str": d_mov_str,
            "d_val_str": d_val_str,
            "desc": desc,
            "deb": deb_num,
            "cred": cred_num,
        })

    # Preparar lote de escrita
    num_cols = max(column_mapping.values()) + 1
    new_rows: list[list[Any]] = []
    ids_gerados: list[str] = []
    ids_reutilizados: list[str] = []
    duplicados_count = 0
    conflitos_count = 0
    validos_count = 0
    revisoes_count = 0

    current_ids_pool = list(existing_ids)

    for item in movements:
        # Pular linha 5 (cabeçalho vazio) - não inserir na sheet e não gerar ID
        if item.line == 5 and not item.data_movimento.strip():
            continue

        d_mov_formatted = format_date_for_sheet(item.data_movimento, year)
        d_val_formatted = format_date_for_sheet(item.data_valor, year)
        desc_lower = item.descricao.strip().lower()

        deb_val = parse_money(item.debito_eur)
        cred_val = parse_money(item.credito_eur)

        d_mov_dm = extract_day_month(item.data_movimento)
        d_val_dm = extract_day_month(item.data_valor)

        matched_existing_id = None
        matched_duplicate = False
        matched_conflict = False

        for ex in existing_rows_list:
            # Deteção de mesma linha por descrição semelhante e montantes ou data
            same_desc = is_same_description(ex["desc"], desc_lower)
            same_date = (ex["d_mov_dm"] == d_mov_dm and d_mov_dm is not None) or (ex["d_mov_str"] == d_mov_formatted) or (ex["d_mov_str"] == item.data_movimento.strip())
            same_amounts = (ex["deb"] == deb_val) and (ex["cred"] == cred_val)

            if (same_desc and same_amounts) or (same_date and same_desc and same_amounts):
                matched_duplicate = True
                matched_existing_id = ex["id"]
                break
            elif same_desc and same_date and not same_amounts:
                matched_conflict = True
                break

        if matched_duplicate:
            duplicados_count += 1
            if matched_existing_id:
                ids_reutilizados.append(matched_existing_id)
            continue

        if matched_conflict:
            conflitos_count += 1
            continue

        new_id = generate_next_id_interno(current_ids_pool, prefix=prefix, digits=digits)
        current_ids_pool.append(new_id)
        ids_gerados.append(new_id)

        desc_clean, pais_code = extract_country_and_clean_desc(item.descricao, item.pais)
        val_orig_str, moeda_code = parse_original_amount_and_currency(item.moeda_original)

        row_cells: list[Any] = [""] * num_cols
        row_cells[column_mapping["data_movimento"]] = d_mov_formatted
        row_cells[column_mapping["data_valor"]] = d_val_formatted
        row_cells[column_mapping["descricao"]] = desc_clean
        row_cells[column_mapping["pais"]] = pais_code
        row_cells[column_mapping["valor_original"]] = val_orig_str
        row_cells[column_mapping["moeda_original"]] = moeda_code
        row_cells[column_mapping["taxa_cambio"]] = item.taxa_cambio
        row_cells[column_mapping["debito_eur"]] = deb_val if deb_val is not None else ""
        row_cells[column_mapping["credito_eur"]] = cred_val if cred_val is not None else ""
        row_cells[column_mapping["id_interno"]] = new_id
        row_cells[column_mapping["finance"]] = ""
        row_cells[column_mapping["status"]] = item.status
        row_cells[column_mapping["motivos_revisao"]] = item.motivos_revisao if item.status == "REVISÃO" else ""

        if item.status == "VÁLIDO":
            validos_count += 1
        else:
            revisoes_count += 1

        new_rows.append(row_cells)

    stats = {
        "sheet_write": False if dry_run else True,
        "dry_run": dry_run,
        "sheet_name": worksheet,
        "movimentos_processados": len(movements),
        "linhas_novas": len(new_rows),
        "linhas_inseridas": 0 if dry_run else len(new_rows),
        "duplicados_ignorados": duplicados_count,
        "conflitos": conflitos_count,
        "validos_inseridos": validos_count if not dry_run else 0,
        "revisoes_inseridas": revisoes_count if not dry_run else 0,
        "ids_gerados": ids_gerados,
        "ids_reutilizados": ids_reutilizados,
    }

    if dry_run or not new_rows:
        saidas_stats = sync_movements_to_saidas(
            sheets_service,
            cfg,
            movements,
            values,
            drive_info,
            service_account_email,
        )
        stats.update(saidas_stats)
        return stats

    start_row_num = None
    for idx in range(1, len(values)):
        r = values[idx]
        col_id = str(r[column_mapping["id_interno"]]).strip() if "id_interno" in column_mapping and len(r) > column_mapping["id_interno"] else ""
        col_d = str(r[column_mapping["data_movimento"]]).strip() if "data_movimento" in column_mapping and len(r) > column_mapping["data_movimento"] else ""
        if not col_id and not col_d:
            start_row_num = idx + 1
            break
    if start_row_num is None:
        start_row_num = len(values) + 1
    try:
        sheets_service.spreadsheets().values().update(
            spreadsheetId=spreadsheet_id,
            range=f"'{worksheet}'!A{start_row_num}",
            valueInputOption="USER_ENTERED",
            body={"values": new_rows},
        ).execute()

        try:
            meta = sheets_service.spreadsheets().get(spreadsheetId=spreadsheet_id).execute()
            sheet_id = next((s["properties"]["sheetId"] for s in meta.get("sheets", []) if s["properties"]["title"] == worksheet), 0)
            req = {
                "requests": [{
                    "repeatCell": {
                        "range": {
                            "sheetId": sheet_id,
                            "startRowIndex": start_row_num - 1,
                            "endRowIndex": start_row_num - 1 + len(new_rows),
                            "startColumnIndex": 0,
                            "endColumnIndex": num_cols,
                        },
                        "cell": {
                            "userEnteredFormat": {
                                "textFormat": {"bold": False},
                            }
                        },
                        "fields": "userEnteredFormat.backgroundColor,userEnteredFormat.backgroundColorStyle,userEnteredFormat.textFormat.bold",
                    }
                }]
            }
            sheets_service.spreadsheets().batchUpdate(spreadsheetId=spreadsheet_id, body=req).execute()
        except Exception:
            pass
    except HttpError as err:
        if getattr(err, "resp", None) and getattr(err.resp, "status", None) == 403:
            raise RuntimeError(
                f"Partilhe o ficheiro AppTesouraria com:\n{service_account_email}\natribuindo a permissão Editor."
            ) from err
        raise RuntimeError(f"Erro ao escrever no Google Sheets: {err}") from err

    post_res = sheets_service.spreadsheets().values().get(
        spreadsheetId=spreadsheet_id,
        range=f"'{worksheet}'!A1:Z5000",
    ).execute()
    post_values = post_res.get("values", [])
    post_ids = [str(r[id_col]).strip() for r in post_values[1:] if len(r) > id_col and str(r[id_col]).strip()]
    if len(post_ids) != len(set(post_ids)):
        raise RuntimeError("CRÍTICO: Foram detetados IDs duplicados na Sheet após a escrita!")

    saidas_stats = sync_movements_to_saidas(
        sheets_service,
        cfg,
        movements,
        post_values if 'post_values' in locals() else None,
        drive_info,
        service_account_email,
    )
    stats.update(saidas_stats)

    return stats


def load_constantes_details_map(sheets_service: Any, spreadsheet_id: str, worksheet: str = "CONSTANTES") -> dict[str, dict[str, str]]:
    if not sheets_service or not spreadsheet_id:
        return {}
    try:
        response = sheets_service.spreadsheets().values().get(
            spreadsheetId=spreadsheet_id,
            range=f"'{worksheet}'!A1:Z5000",
        ).execute()
        rows = response.get("values", [])
        if not rows:
            return {}
        header = [str(h).strip().lower() for h in rows[0]]
        texto_col = next((idx for idx, c in enumerate(header) if "texto" in c), 1)
        plano_col = next((idx for idx, c in enumerate(header) if "plano de conta" in c), None)
        centro_col = next((idx for idx, c in enumerate(header) if "centro de custo" in c), None)
        desc_soma_col = next((idx for idx, c in enumerate(header) if "descrição soma" in c or "descricao soma" in c), None)
        tipo_col = next((idx for idx, c in enumerate(header) if c == "tipo"), None)
        forma_col = next((idx for idx, c in enumerate(header) if "forma de pagamento" in c), None)
        caixa_col = next((idx for idx, c in enumerate(header) if c == "caixa"), None)

        details = {}
        for r in rows[1:]:
            if len(r) > texto_col and str(r[texto_col]).strip():
                key = str(r[texto_col]).strip().upper()
                details[key] = {
                    "plano_de_conta": str(r[plano_col]).strip() if plano_col is not None and len(r) > plano_col else "",
                    "centro_de_custo": str(r[centro_col]).strip() if centro_col is not None and len(r) > centro_col else "",
                    "descricao_soma": str(r[desc_soma_col]).strip() if desc_soma_col is not None and len(r) > desc_soma_col else "",
                    "tipo": str(r[tipo_col]).strip() if tipo_col is not None and len(r) > tipo_col else "",
                    "forma_de_pagamento": str(r[forma_col]).strip() if forma_col is not None and len(r) > forma_col else "",
                    "caixa": str(r[caixa_col]).strip() if caixa_col is not None and len(r) > caixa_col else "",
                }
        return details
    except Exception:
        return {}


def sync_movements_to_saidas(
    sheets_service: Any,
    cfg: dict[str, Any],
    movements: list[Movement],
    cartao_values: list[list[Any]] | None,
    drive_image_info: dict[str, Any] | None,
    service_account_email: str,
) -> dict[str, Any]:
    sheets_cfg = cfg.get("google_sheets", {})
    if not sheets_cfg.get("enabled", False):
        return {"saidas_write": False}

    spreadsheet_id = str(sheets_cfg.get("spreadsheet_id", "")).strip()
    worksheet = str(sheets_cfg.get("saidas_worksheet", "SAÍDAS")).strip()
    dry_run = bool(sheets_cfg.get("dry_run", True))
    prefix = str(sheets_cfg.get("saidas_id_prefix", "SAI")).strip()
    digits = int(sheets_cfg.get("id_digits", 10))

    if not spreadsheet_id:
        raise ValueError("ID da folha de cálculo 'google_sheets.spreadsheet_id' não configurado.")

    constantes_details = load_constantes_details_map(sheets_service, spreadsheet_id)

    try:
        response = sheets_service.spreadsheets().values().get(
            spreadsheetId=spreadsheet_id,
            range=f"'{worksheet}'!A1:Z5000",
        ).execute()
    except Exception as exc:
        raise RuntimeError(f"Erro ao aceder à sheet '{worksheet}': {exc}") from exc

    values = response.get("values", [])
    if not values:
        raise ValueError(f"A sheet '{worksheet}' está totalmente vazia ou não contém cabeçalho.")

    header_row = values[0]
    header_normalized = [str(h).strip().lower() for h in header_row]
    column_mapping: dict[str, int] = {norm: idx for idx, norm in enumerate(header_normalized)}

    existing_ids: list[str] = []
    existing_rows_list: list[dict[str, Any]] = []

    id_col = column_mapping.get("id_interno")
    data_col = column_mapping.get("data")
    desc_col = column_mapping.get("descrição da compra")
    val_col = column_mapping.get("valor da compra")

    year = int(cfg.get("validation", {}).get("year", 2026))

    for row in values[1:]:
        val_id = str(row[id_col]).strip() if id_col is not None and len(row) > id_col else ""
        if val_id:
            existing_ids.append(val_id)

        d_str = str(row[data_col]).strip() if data_col is not None and len(row) > data_col else ""
        desc = str(row[desc_col]).strip().lower() if desc_col is not None and len(row) > desc_col else ""
        val = parse_money(str(row[val_col])) if val_col is not None and len(row) > val_col and str(row[val_col]).strip() else None

        existing_rows_list.append({
            "id": val_id,
            "data_str": d_str,
            "data_dm": extract_day_month(d_str),
            "desc": desc,
            "val": val,
        })

    num_cols = max(column_mapping.values()) + 1
    new_rows: list[list[Any]] = []
    ids_gerados: list[str] = []
    duplicados_count = 0

    current_ids_pool = list(existing_ids)
    image_name = drive_image_info.get("name", "") if drive_image_info else ""
    now_str = datetime.now().strftime("%d/%m/%Y %H:%M:%S")

    for item in movements:
        d_mov_formatted = format_date_for_sheet(item.data_movimento, year)
        desc_lower = item.descricao.strip().lower()
        deb_val = parse_money(item.debito_eur)

        if deb_val is None:
            continue

        d_mov_dm = extract_day_month(item.data_movimento)
        matched_duplicate = False

        for ex in existing_rows_list:
            same_desc = is_same_description(ex["desc"], desc_lower)
            same_date = (ex["data_dm"] == d_mov_dm and d_mov_dm is not None) or (ex["data_str"] == d_mov_formatted)
            same_amount = (ex["val"] == deb_val)

            if same_desc and same_date and same_amount:
                matched_duplicate = True
                break

        if matched_duplicate:
            duplicados_count += 1
            continue

        new_id = generate_next_id_interno(current_ids_pool, prefix=prefix, digits=digits)
        current_ids_pool.append(new_id)
        ids_gerados.append(new_id)

        desc_clean, pais_code = extract_country_and_clean_desc(item.descricao, item.pais)

        # Procurar correspondência na sheet CONSTANTES
        c_info = None
        desc_clean_upper = desc_clean.upper().strip()
        for t_key, info in constantes_details.items():
            if t_key and (t_key in desc_clean_upper or desc_clean_upper in t_key or is_same_description(t_key, desc_clean_upper)):
                c_info = info
                break

        forma_pag = (c_info["forma_de_pagamento"] if c_info and c_info["forma_de_pagamento"] else "") or "CARTÃO DE CRÉDITO"
        tipo_val = (c_info["tipo"] if c_info and c_info["tipo"] else "") or "PAGAMENTO"
        desc_compra = (c_info["descricao_soma"] if c_info and c_info["descricao_soma"] else "") or (f"{desc_clean} {pais_code}".strip() if pais_code else desc_clean)
        plano_conta = c_info["plano_de_conta"] if c_info and c_info["plano_de_conta"] else ""
        centro_custo = (c_info["centro_de_custo"] if c_info and c_info["centro_de_custo"] else "") or "PADRÃO"
        caixa_val = (c_info["caixa"] if c_info and c_info["caixa"] else "") or "CARTÃO DE CRÉDITO"

        row_cells: list[Any] = [""] * num_cols

        if "id_interno" in column_mapping:
            row_cells[column_mapping["id_interno"]] = new_id
        if "forma de pagamento" in column_mapping:
            row_cells[column_mapping["forma de pagamento"]] = forma_pag
        if "data" in column_mapping:
            row_cells[column_mapping["data"]] = d_mov_formatted
        if "data valor" in column_mapping:
            row_cells[column_mapping["data valor"]] = format_date_for_sheet(item.data_valor, year)
        if "tipo" in column_mapping:
            row_cells[column_mapping["tipo"]] = tipo_val

        # IMPORTANTE: "DOC. SOMA" e "STATUS PROCESSAMENTO" NÃO são preenchidos por instrução explícita do utilizador
        if "doc. soma" in column_mapping:
            row_cells[column_mapping["doc. soma"]] = ""
        if "status processamento" in column_mapping:
            row_cells[column_mapping["status processamento"]] = ""

        if "nº recibo" in column_mapping:
            row_cells[column_mapping["nº recibo"]] = desc_clean
        if "valor da compra" in column_mapping:
            row_cells[column_mapping["valor da compra"]] = deb_val
        if "descrição da compra" in column_mapping:
            row_cells[column_mapping["descrição da compra"]] = desc_compra
        if "nome do fornecedor" in column_mapping:
            row_cells[column_mapping["nome do fornecedor"]] = desc_clean
        if "plano de conta" in column_mapping:
            row_cells[column_mapping["plano de conta"]] = plano_conta
        if "centro de custo" in column_mapping:
            row_cells[column_mapping["centro de custo"]] = centro_custo
        if "caixa" in column_mapping:
            row_cells[column_mapping["caixa"]] = caixa_val
        if "imagem do recibo" in column_mapping:
            row_cells[column_mapping["imagem do recibo"]] = image_name
        if "status da tesouraria" in column_mapping:
            row_cells[column_mapping["status da tesouraria"]] = "Concluído" if item.status == "VÁLIDO" else "Pendente"
        if "finance" in column_mapping:
            row_cells[column_mapping["finance"]] = "Enviado"
        if "iduser" in column_mapping:
            row_cells[column_mapping["iduser"]] = service_account_email
        if "timestamp" in column_mapping:
            row_cells[column_mapping["timestamp"]] = now_str

        new_rows.append(row_cells)

    stats = {
        "saidas_write": False if dry_run else True,
        "saidas_sheet": worksheet,
        "saidas_linhas_novas": len(new_rows),
        "saidas_duplicados_ignorados": duplicados_count,
        "saidas_ids_gerados": ids_gerados,
    }

    if dry_run or not new_rows:
        return stats

    start_row_num = None
    for idx in range(1, len(values)):
        r = values[idx]
        col_id = str(r[column_mapping["id_interno"]]).strip() if "id_interno" in column_mapping and len(r) > column_mapping["id_interno"] else ""
        col_d = str(r[column_mapping["data"]]).strip() if "data" in column_mapping and len(r) > column_mapping["data"] else ""
        if not col_id and not col_d:
            start_row_num = idx + 1
            break
    if start_row_num is None:
        start_row_num = len(values) + 1
    sheets_service.spreadsheets().values().update(
        spreadsheetId=spreadsheet_id,
        range=f"'{worksheet}'!A{start_row_num}",
        valueInputOption="USER_ENTERED",
        body={"values": new_rows},
    ).execute()

    return stats


def load_constantes_trusted_texts(sheets_service: Any, spreadsheet_id: str, worksheet: str = "CONSTANTES") -> set[str]:
    if not sheets_service or not spreadsheet_id:
        return set()
    try:
        response = sheets_service.spreadsheets().values().get(
            spreadsheetId=spreadsheet_id,
            range=f"'{worksheet}'!A1:Z5000",
        ).execute()
        rows = response.get("values", [])
        if not rows:
            return set()
        header = [str(h).strip().lower() for h in rows[0]]
        texto_col = next((idx for idx, c in enumerate(header) if "texto" in c), 1)
        return {str(r[texto_col]).strip() for r in rows[1:] if len(r) > texto_col and str(r[texto_col]).strip()}
    except Exception:
        return set()




def get_next_unprocessed_extrato_row(sheets_service: Any, spreadsheet_id: str, worksheet: str = "EXTRATO_CARTÃO") -> tuple[int, str, str] | None:
    if not sheets_service or not spreadsheet_id:
        return None
    try:
        response = sheets_service.spreadsheets().values().get(
            spreadsheetId=spreadsheet_id,
            range=f"'{worksheet}'!A1:Z500",
        ).execute()
        rows = response.get("values", [])
        if not rows:
            return None

        # Análise do cabeçalho
        header = [str(h).strip().lower() for h in rows[0]]
        header_original = [str(h).strip() for h in rows[0]]
        img_col = next((idx for idx, c in enumerate(header) if "imagem" in c), 1)
        status_col = next((idx for idx, c in enumerate(header) if "status" in c), 2)
        nr_col = next((idx for idx, c in enumerate(header) if "extrato" in c), 0)

        # Debug: Mostrar cabeçalho e colunas detectadas
        print(f"\n[EXTRATO_CARTÃO] 📋 ANÁLISE DA SHEET:")
        print(f"{'='*85}")
        print(f"Cabeçalho: {header_original}")
        print(f"\nColunas Detectadas:")
        print(f"  • IMAGEM  (índice {img_col}): '{header_original[img_col] if img_col < len(header_original) else 'N/A'}'")
        print(f"  • STATUS  (índice {status_col}): '{header_original[status_col] if status_col < len(header_original) else 'N/A'}'")
        print(f"  • EXTRATO (índice {nr_col}): '{header_original[nr_col] if nr_col < len(header_original) else 'N/A'}'")
        print(f"\n{'─'*85}")

        # Coletar linhas disponíveis
        available_lines = []
        for idx in range(1, len(rows)):
            r = rows[idx]
            status_val = str(r[status_col]).strip() if len(r) > status_col else ""
            img_val = str(r[img_col]).strip() if len(r) > img_col else ""
            nr_val = str(r[nr_col]).strip() if len(r) > nr_col else ""

            if not status_val and img_val:
                available_lines.append((idx + 1, nr_val, img_val))

        # Mostrar linhas disponíveis
        print(f"Linhas Disponíveis para Processamento (STATUS vazio):\n")
        if available_lines:
            print(f"{'Linha':<8} {'Extrato':<20} {'Imagem':<50}")
            print(f"{'-'*85}")
            for line_num, nr, img in available_lines:
                print(f"{line_num:<8} {nr:<20} {img:<50}")
            print(f"{'-'*85}")
            print(f"Total: {len(available_lines)} linha(s) disponível(is)\n")
        else:
            print(f"⚠️  Nenhuma linha disponível para processamento!\n")

        print(f"{'='*85}\n")

        # Retornar primeira disponível
        if available_lines:
            return available_lines[0]
        return None
    except Exception as e:
        print(f"[ERRO] Falha ao ler sheet EXTRATO_CARTÃO: {e}")
        return None


def mark_extrato_row_completed(sheets_service: Any, spreadsheet_id: str, row_idx: int, worksheet: str = "EXTRATO_CARTÃO") -> None:
    if not sheets_service or not spreadsheet_id or not row_idx:
        return
    try:
        now_str = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
        response = sheets_service.spreadsheets().values().get(
            spreadsheetId=spreadsheet_id,
            range=f"'{worksheet}'!A1:Z1",
        ).execute()
        rows = response.get("values", [])
        if not rows:
            return
        header = [str(h).strip().lower() for h in rows[0]]
        status_col = next((idx for idx, c in enumerate(header) if "status" in c), 2)
        ts_col = next((idx for idx, c in enumerate(header) if "timestamp" in c), 3)

        status_col_letter = chr(65 + status_col)
        ts_col_letter = chr(65 + ts_col)

        sheets_service.spreadsheets().values().update(
            spreadsheetId=spreadsheet_id,
            range=f"'{worksheet}'!{status_col_letter}{row_idx}",
            valueInputOption="USER_ENTERED",
            body={"values": [["Concluido"]]},
        ).execute()

        sheets_service.spreadsheets().values().update(
            spreadsheetId=spreadsheet_id,
            range=f"'{worksheet}'!{ts_col_letter}{row_idx}",
            valueInputOption="USER_ENTERED",
            body={"values": [[now_str]]},
        ).execute()
    except Exception:
        pass


def main() -> int:
    parser = argparse.ArgumentParser(description="OCR seguro de extratos de cartão")
    parser.add_argument("imagem", type=Path, nargs="?", help="Imagem local (opcional; por defeito procura no Drive)")
    parser.add_argument("--config", type=Path, default=Path(__file__).with_name("config.yaml"))
    args = parser.parse_args()

    cfg = load_config(args.config)
    out_dir = Path(cfg.get("output", {}).get("directory", "output"))
    out_dir.mkdir(parents=True, exist_ok=True)

    scopes = [
        "https://www.googleapis.com/auth/drive.readonly",
        "https://www.googleapis.com/auth/cloud-platform",
        "https://www.googleapis.com/auth/spreadsheets",
    ]
    credentials = load_google_credentials(cfg, scopes)
    service_account_email = getattr(credentials, "service_account_email", "conta-de-serviço")

    drive_info = None
    with tempfile.TemporaryDirectory(prefix="soma_ocr_") as temp_dir:
        if args.imagem:
            source_image = args.imagem
            source_label = str(source_image)
        sheets_service = create_sheets_service(credentials)
        spreadsheet_id = str(cfg.get("google_sheets", {}).get("spreadsheet_id", "")).strip()
        extrato_info = get_next_unprocessed_extrato_row(sheets_service, spreadsheet_id)

        if not args.imagem and extrato_info:
            extrato_row_idx, nr_extrato, img_rel_path = extrato_info
            img_filename = Path(img_rel_path).name
            print(f"[EXTRATO_CARTÃO] A processar extrato pendente da linha {extrato_row_idx}: {img_filename}")
            drive_cfg = cfg.get("drive") or cfg.get("input", {}).get("drive", {})
            drive_cfg_temp = dict(drive_cfg)
            drive_cfg_temp["filename"] = img_filename
            drive_service = create_drive_service(credentials)
            drive_info = locate_drive_image(drive_service, drive_cfg_temp, service_account_email)
            source_image = Path(temp_dir) / img_filename
            download_drive_image(drive_service, drive_info["id"], source_image)
            source_label = f"Google Drive: {drive_info['name']} (ID {drive_info['id']})"
        elif not args.imagem:
            drive_cfg = cfg.get("drive") or cfg.get("input", {}).get("drive", {})
            filename = str(drive_cfg.get("filename", "")).strip()
            folder_name = str(drive_cfg.get("folder_name", "EXTRATO_CARTÃO_Images")).strip()
            print(f"[DRIVE] A localizar no Google Drive: {folder_name} / {filename}")

            drive_service = create_drive_service(credentials)
            drive_info = locate_drive_image(drive_service, drive_cfg, service_account_email)
            print(f"[DRIVE] Imagem localizada: {drive_info['name']} | ID: {drive_info['id']}")

            source_image = Path(temp_dir) / filename
            download_drive_image(drive_service, drive_info["id"], source_image)
            source_label = f"Google Drive: {drive_info['name']} (ID {drive_info['id']})"
            print(f"[DRIVE] Imagem descarregada com sucesso.")

        variants = preprocess(source_image, cfg.get("preprocessing", {}), out_dir)
        ocr_image = variants[0]
        result = vision_request(ocr_image, cfg, credentials)
        words = extract_words(result)
        image = cv2.imread(str(ocr_image))
        if image is None:
            raise ValueError(f"Não foi possível carregar a imagem pré-processada: {ocr_image}")

        trusted_texts = load_constantes_trusted_texts(sheets_service, spreadsheet_id)

        rows = group_rows(words, image.shape[1], image.shape[0], cfg.get("table", {}))
        movements = build_movements(rows, image.shape[1], cfg, trusted_texts=trusted_texts)

        # Aprendizado de padrões de comerciantes (Fase 3)
        learner = MerchantPatternLearner()
        valid_movements = [m for m in movements if m.status == "VÁLIDO"]
        for mov in valid_movements:
            learner.learn_from_movement(mov)
        merchant_patterns_summary = learner.get_statistics()

        metadata = {
            "imagem": source_label,
            "drive": drive_info,
            "imagem_ocr": str(ocr_image),
            "palavras": len(words),
            "linhas": len(movements),
            "validas": sum(x.status == "VÁLIDO" for x in movements),
            "revisao": sum(x.status == "REVISÃO" for x in movements),
            "sheet_write": False,
            "merchant_patterns": merchant_patterns_summary,
        }

        sheets_stats = sync_movements_to_sheets(sheets_service, cfg, movements, drive_info, service_account_email)
        metadata.update(sheets_stats)

        # Geração de métricas de qualidade (relatório)
        quality_metrics = generate_ocr_quality_metrics(movements, cfg)
        metrics_report = format_metrics_report(quality_metrics)
        metadata["quality_metrics"] = quality_metrics

        dry_run = bool(cfg.get("google_sheets", {}).get("dry_run", True))
        if extrato_info and not dry_run:
            mark_extrato_row_completed(sheets_service, spreadsheet_id, extrato_info[0])

        save_outputs(movements, out_dir, metadata)
        annotate(ocr_image, words, movements, out_dir / "03_diagnostico.png")

        # Guardar relatório de qualidade
        (out_dir / "relatorio_qualidade.txt").write_text(metrics_report, encoding="utf-8")

    print(json.dumps(metadata, ensure_ascii=False, indent=2))
    print(f"Resultados guardados em: {out_dir.resolve()}")
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except Exception as exc:
        print(f"ERRO: {exc}", file=sys.stderr)
        raise SystemExit(1)
